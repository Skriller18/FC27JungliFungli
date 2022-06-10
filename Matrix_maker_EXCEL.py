
#importing cv2 so as to read the map image to find out the different routes
#These routes are later converted into matrix of 1,0 which will be used for getting shortest route using A* algorithm
#A nested list is used as a matrix
#Making a black pixel as 1 and other color as 0
import cv2
import numpy as np
import openpyxl
from openpyxl import load_workbook

#import xlsxw
#path_output="C:\\User\\Om\\OneDrive\\Desktop\\Code\\matrixfile.xlsx"
#wboutput=openpyxl.load_workbook(filename=path_output)
#soutput=wboutput.active

#path_input="C:\\Users\\Om\\OneDrive\\Desktop\\BECIL_code\\dummy_input2.xlsx"
path_output="C:\\Users\\Om\\OneDrive\\Desktop\\BECIL_code\\dummy_output6.xlsx"

#wbinput=openpyxl.load_workbook(filename=path_input)
wboutput=openpyxl.load_workbook(filename=path_output)
#sinput=wbinput.active
soutput=wboutput.active
try:
    # Reading the image using imread() function
    image = cv2.imread('black_image.jpg')

      
    # Extracting the height and width of an image
    h = image.shape[0]
    w = image.shape[1]


    # Displaying the height and width
    print("Height = {},  Width = {}".format(h, w))
    #Making nested list
    matrix = []
      
    for i in range(10):
          
        # Append an empty sublist inside the list
        #matrix.append([])
        line=[]
          
        for j in range(5):
            (B, G, R) = image[j, i]
            flag = True
            if(R<=10) and (B<=10) and (G<=10):
                #matrix[i][j] = 1
                #matrix[i].append(j)
                c=soutput.cell(row=i+1,column=j+1)
                c.value=1
                line.append(1)
            else:
                c=soutput.cell(row=i+1,column=j+1)
                c.value=0
                line.append(0)
        matrix.append(line)
              
    print(matrix)

    #np.set_printoptions(threshold=np.inf)
    #print(matrix)
    cv2.imshow("Image",image)
    cv2.waitKey()
    cv2.destroy()

except AttributeError:
    print("Image closed")
print("Succesfully added")
wboutput.save(filename="C:\\Users\\Om\\OneDrive\\Desktop\\BECIL_code\\matrixfile.xlsx")


path_input="C:\\Users\\Om\\OneDrive\\Desktop\\BECIL_code\\matrixfile.xlsx"
wbinput=openpyxl.load_workbook(filename=path_input)
sinput=wbinput.active
int1=sinput.cell(row=4,column=3)
i1=int(int1.value)
int2=sinput.cell(row=9,column=3)
i2=int(int2.value)
print(i1,i2)
